"""Export des donnees brutes vers un fichier .xlsx pour l'Account Manager.

Un onglet par source. Aucune analyse LLM, aucun rapport genere: on organise
juste les JSON de `raw/` en tableaux exploitables comme matiere premiere pour
un audit manuel."""
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.storage import ClientWorkspace

CELL_TEXT_MAX = 4000  # Excel limite a 32767 mais >4k tue la lisibilite
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(vertical="center", horizontal="left")


def _clip(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        v = str(v)
    if isinstance(v, str) and len(v) > CELL_TEXT_MAX:
        return v[:CELL_TEXT_MAX] + " ...[tronque]"
    return v


def _get(d: dict, *keys, default=""):
    for k in keys:
        if k in d and d[k] not in (None, ""):
            return d[k]
    return default


def _write_sheet(wb: Workbook, title: str, headers: list[str],
                 rows: Iterable[list[Any]]) -> int:
    ws = wb.create_sheet(title[:31])  # Excel limite le nom d'onglet a 31 chars
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    n = 0
    for row in rows:
        ws.append([_clip(c) for c in row])
        n += 1
    ws.freeze_panes = "A2"
    # largeurs indicatives (l'utilisateur ajustera si besoin)
    widths = [min(60, max(12, len(h) + 4)) for h in headers]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return n


def _reviews_rows(items: list[dict]) -> Iterable[list]:
    for r in items or []:
        if not isinstance(r, dict):
            continue
        if r.get("limit_reached") or r.get("error"):
            continue  # skip les rows d'erreur laissees par les collecteurs
        source = r.get("_source") or r.get("reviewOrigin") or ""
        origin = r.get("_source_domain") or r.get("_source_query") or ""
        date_v = _get(r, "publishedAtDate", "dateOfExperience", "createdAt",
                      "date", "publishAt")
        rating = _get(r, "stars", "rating", "starRating", "score")
        author = _get(r, "name", "author", "reviewerName", "consumerName")
        text = _get(r, "text", "textTranslated", "review", "body", "content")
        title = _get(r, "title", "reviewTitle")
        url = _get(r, "reviewUrl", "url", "permalink")
        verified = _get(r, "verified", "isVerified")
        yield [source, origin, date_v, rating, author, title, text, url, verified]


def _reddit_rows(items: list[dict]) -> Iterable[list]:
    for entry in items or []:
        if not isinstance(entry, dict):
            continue
        query = entry.get("query", "")
        post = entry.get("post") or {}
        yield [
            query, post.get("subreddit", ""), post.get("createdAt", ""),
            "post", post.get("title", ""), post.get("selftext", ""),
            post.get("score"), post.get("numComments"),
            post.get("author", ""),
            post.get("url") or post.get("permalink", ""),
        ]
        for c in entry.get("comments") or []:
            if not isinstance(c, dict):
                continue
            yield [
                query, post.get("subreddit", ""), c.get("createdAt", ""),
                "comment", "", c.get("body", ""),
                c.get("score"), None,
                c.get("author", ""),
                c.get("permalink", ""),
            ]


def _youtube_rows(items: list[dict]) -> Iterable[list]:
    for v in items or []:
        if not isinstance(v, dict) or v.get("error"):
            continue
        comments = v.get("comments") or []
        top = " | ".join(
            (c.get("text", "") if isinstance(c, dict) else str(c))
            for c in comments[:10]
        )
        yield [
            v.get("_source_query", ""), v.get("title", ""),
            v.get("channel") or v.get("author", ""),
            v.get("views") or v.get("view_count", ""),
            v.get("published") or v.get("publish_time", ""),
            v.get("url", ""),
            v.get("transcript", ""),
            top,
            len(comments),
        ]


def _fb_ads_rows(items: list[dict]) -> Iterable[list]:
    for a in items or []:
        if not isinstance(a, dict):
            continue
        snap = a.get("snapshot") or {}
        body = snap.get("body") or {}
        body_text = body.get("text", "") if isinstance(body, dict) else str(body)
        cards = snap.get("cards") or []
        card_headlines = " | ".join(
            (c.get("title", "") if isinstance(c, dict) else "")
            for c in cards
        )
        yield [
            snap.get("page_name", ""), snap.get("page_id", ""),
            a.get("ad_archive_id", ""),
            a.get("start_date") or a.get("startDate", ""),
            a.get("end_date") or a.get("endDate", ""),
            a.get("is_active") if "is_active" in a else a.get("isActive", ""),
            snap.get("cta_text", ""), snap.get("link_url", ""),
            body_text, card_headlines,
            snap.get("page_profile_uri", ""),
        ]


def _trends_rows(data: dict) -> Iterable[list]:
    if not isinstance(data, dict):
        return
    iot = data.get("interest_over_time") or {}
    for kw, series in iot.items():
        if not isinstance(series, dict):
            continue
        for date_v, val in series.items():
            yield [kw, str(date_v), val]


def _industry_us_rows(data: dict) -> Iterable[list]:
    if not isinstance(data, dict):
        return
    cat = data.get("category", "")
    for pt in data.get("monthly_series") or []:
        yield [cat, pt.get("date", ""), pt.get("value", "")]


def _industry_ca_rows(data: dict) -> Iterable[list]:
    if not isinstance(data, dict):
        return
    cat = data.get("category", "")
    for region, block in (data.get("retail_sales") or {}).items():
        if not isinstance(block, dict):
            continue
        for pt in block.get("monthly_series") or []:
            yield [cat, region, pt.get("date", ""), pt.get("value", "")]


def _sec_rows(data: dict) -> Iterable[list]:
    if not isinstance(data, dict):
        return
    for f in data.get("filings") or []:
        yield [
            f.get("company", ""), f.get("cik", ""), f.get("form", ""),
            f.get("filed", ""), f.get("period_ending", ""),
            ", ".join(f.get("biz_locations") or []),
            f.get("accession", ""),
        ]


def _pages_rows(items: list[dict]) -> Iterable[list]:
    for p in items or []:
        if not isinstance(p, dict):
            continue
        yield [p.get("url", ""), p.get("method", ""),
               p.get("error") or "", p.get("text", "")]


def _summary_rows(ws_root: Path, data_score: dict, onboarding: dict,
                  plan: dict) -> Iterable[list]:
    yield ["client_slug", onboarding.get("client_slug", "")]
    yield ["client_name", onboarding.get("client_name", "")]
    yield ["site", onboarding.get("site", "") or onboarding.get("website", "")]
    yield ["audit_id", ws_root.name]
    yield ["generated_at", datetime.now().isoformat(timespec="seconds")]
    yield ["", ""]
    yield ["--- Compteurs de collecte ---", ""]
    for k, v in (data_score or {}).items():
        yield [k, v]
    yield ["", ""]
    yield ["--- Plan de collecte ---", ""]
    for k, v in (plan or {}).items():
        yield [k, ", ".join(v) if isinstance(v, list) else str(v)]


def render_excel(ws: ClientWorkspace, client_name: str) -> str:
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    onboarding = ws.read_json(ws.root / "onboarding.json") or {}
    plan = ws.read_json(ws.root / "collection-plan.json") or {}
    data_score = ws.read_json(ws.root / "data-score.json") or {}

    _write_sheet(wb, "_Summary", ["cle", "valeur"],
                 _summary_rows(ws.root, data_score, onboarding, plan))

    reviews = ws.read_json(ws.raw("trustpilot.json")) or []
    _write_sheet(wb, "Reviews",
                 ["source", "origine", "date", "note", "auteur", "titre",
                  "texte", "url", "verifie"],
                 _reviews_rows(reviews))

    reddit = ws.read_json(ws.raw("reddit.json")) or []
    _write_sheet(wb, "Reddit",
                 ["query", "subreddit", "date", "type", "titre", "texte",
                  "score", "num_comments", "auteur", "url"],
                 _reddit_rows(reddit))

    youtube = ws.read_json(ws.raw("youtube.json")) or []
    _write_sheet(wb, "YouTube",
                 ["query", "titre_video", "chaine", "vues", "publie",
                  "url", "transcript", "top_commentaires", "n_commentaires"],
                 _youtube_rows(youtube))

    ads = ws.read_json(ws.raw("fb_ads.json")) or []
    _write_sheet(wb, "FB_Ads",
                 ["page", "page_id", "ad_id", "debut", "fin", "actif",
                  "cta", "link_url", "corps", "carousel_titres", "page_url"],
                 _fb_ads_rows(ads))

    trends = ws.read_json(ws.raw("google_trends.json")) or {}
    _write_sheet(wb, "Google_Trends",
                 ["mot_cle", "date", "interet_0_100"],
                 _trends_rows(trends))

    ind_us = ws.read_json(ws.raw("industry_us.json")) or {}
    _write_sheet(wb, "Industry_US",
                 ["categorie", "date", "ventes_M_USD"],
                 _industry_us_rows(ind_us))

    ind_ca = ws.read_json(ws.raw("industry_ca.json")) or {}
    _write_sheet(wb, "Industry_CA",
                 ["categorie", "region", "date", "ventes_k_CAD"],
                 _industry_ca_rows(ind_ca))

    sec = ws.read_json(ws.raw("sec_filings.json")) or {}
    _write_sheet(wb, "SEC_Filings",
                 ["entreprise", "cik", "formulaire", "depose_le",
                  "periode_fin", "localisations", "accession"],
                 _sec_rows(sec))

    client_pages = ws.read_json(ws.raw("client_pages.json")) or []
    _write_sheet(wb, "Client_Pages",
                 ["url", "methode", "erreur", "texte_extrait"],
                 _pages_rows(client_pages))

    comp_pages = ws.read_json(ws.raw("competitor_pages.json")) or []
    _write_sheet(wb, "Competitor_Pages",
                 ["url", "methode", "erreur", "texte_extrait"],
                 _pages_rows(comp_pages))

    articles = ws.read_json(ws.raw("industry_articles.json")) or []
    if articles:
        _write_sheet(wb, "Industry_Articles",
                     ["url", "methode", "erreur", "texte_extrait"],
                     _pages_rows(articles))

    semrush = ws.read_json(ws.raw("semrush.json")) or {}
    if semrush:
        _write_sheet(wb, "Semrush", ["cle", "valeur"],
                     ([k, str(v)] for k, v in semrush.items()))

    out = ws.report("audit_data.xlsx")
    wb.save(str(out))
    return str(out)
